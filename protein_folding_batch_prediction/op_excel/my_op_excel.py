'''create by jji on 2021-9-23'''


from openpyxl import load_workbook

file_source = '达安眼科基因检测-疾病候选基因列表210923.xlsx'

work_book = load_workbook(file_source)
work_sheet0 = work_book.worksheets[0]


for index, row in enumerate(work_sheet0.rows):
    if index == 0:
        continue

    code1 = row[0].value
    print(code1)

    if code1 != None:
        code_current = code1

    disease_name = row[1].value
    if disease_name != None:
        code_disease_name = code1

    print(disease_name)

    work_sheet0.cell(row=index + 1, column=6).value = code_disease_name